import os
import re
import asyncio
import psycopg2
from openpyxl import load_workbook

from flask import Flask, request
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)


# =========================================================
# НАСТРОЙКИ
# =========================================================

BOT_TOKEN = os.environ["BOT_TOKEN"]

# Два постоянных администратора.
# На Render задаются переменные ADMIN_ID_1 и ADMIN_ID_2.
ADMIN_ID_1 = int(os.environ["ADMIN_ID_1"])
ADMIN_ID_2 = int(os.environ["ADMIN_ID_2"])
ADMIN_IDS = {ADMIN_ID_1, ADMIN_ID_2}

WEBHOOK_SECRET = "sklad-kustikov-2026-secret-8472"

EXCEL_FILE = "flowers.xlsx"

app = Flask(__name__)

telegram_app = (
    Application.builder()
    .token(BOT_TOKEN)
    .updater(None)
    .build()
)


# =========================================================
# DATABASE
# =========================================================

def get_database_url():
    url = os.environ.get("DATABASE_URL")

    if url:
        return url.strip()

    secret_files = [
        "/etc/secrets/DATABASE_URL",
        "/etc/secrets/database_url",
        "/etc/secrets/neon_url",
    ]

    for filename in secret_files:
        if os.path.exists(filename):
            with open(filename, "r", encoding="utf-8") as f:
                url = f.read().strip()

            if url:
                return url

    raise RuntimeError(
        "DATABASE_URL не найдена."
    )


def get_db():
    return psycopg2.connect(
        get_database_url(),
        connect_timeout=10
    )


# =========================================================
# NORMALIZE
# =========================================================

def normalize(text):
    return " ".join(
        str(text)
        .lower()
        .replace("ё", "е")
        .split()
    )


# =========================================================
# PLAYER / COMMAND STATE
# =========================================================

PLAYER_ID_RE = re.compile(r"^[A-Za-z0-9]+$")

# Состояния именно команд /at и /add.
# Ключ = (telegram_user_id, chat_id), поэтому личка и группа
# никогда не продолжают один и тот же диалог.
COMMAND_STATES = {}


def is_valid_player_id(value):
    return bool(PLAYER_ID_RE.fullmatch(str(value).strip()))


def command_state_key(update):
    return (
        update.effective_user.id,
        update.effective_chat.id
    )


def clear_command_state(update):
    COMMAND_STATES.pop(command_state_key(update), None)


def get_command_state(update):
    return COMMAND_STATES.get(command_state_key(update))


# =========================================================
# DATABASE INIT
# =========================================================

def init_db():
    db = get_db()

    try:
        with db.cursor() as cursor:

            # Старая основная таблица
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS flowers (
                    id SERIAL PRIMARY KEY,
                    flower TEXT NOT NULL,
                    person TEXT NOT NULL
                )
            """)

            cursor.execute("""
                ALTER TABLE flowers
                ADD COLUMN IF NOT EXISTS flower_id TEXT
            """)

            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_flowers_flower
                ON flowers(flower)
            """)

            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_flowers_flower_id
                ON flowers(flower_id)
            """)

            # Постоянные игроки: ID не меняется, ник может меняться.
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS players (
                    player_id TEXT PRIMARY KEY,
                    nickname TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_players_nickname
                ON players(nickname)
            """)

            # Переносим уже существующих игроков из старой таблицы цветов.
            # Берём последнее имя, если исторически у одного ID их было несколько.
            cursor.execute("""
                INSERT INTO players (player_id, nickname)
                SELECT DISTINCT ON (flower_id)
                    flower_id,
                    person
                FROM flowers
                WHERE flower_id IS NOT NULL
                  AND TRIM(flower_id) <> ''
                  AND TRIM(person) <> ''
                ORDER BY flower_id, id DESC
                ON CONFLICT (player_id) DO NOTHING
            """)

            # Администраторы
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS bot_admins (
                    telegram_id BIGINT PRIMARY KEY,
                    role TEXT NOT NULL DEFAULT 'admin',
                    added_by BIGINT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Пользователи, которые писали боту
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS bot_users (
                    telegram_id BIGINT PRIMARY KEY,
                    username TEXT,
                    first_name TEXT,
                    last_name TEXT,
                    chat_id BIGINT,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Настройки бота
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS bot_settings (
                    setting_key TEXT PRIMARY KEY,
                    setting_value TEXT NOT NULL
                )
            """)

            # Журнал действий
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS bot_logs (
                    id SERIAL PRIMARY KEY,
                    admin_id BIGINT,
                    action TEXT,
                    details TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Два постоянных администратора.
            # Оба получают одинаковые права администратора,
            # включая управление другими администраторами.
            for admin_id in ADMIN_IDS:
                cursor.execute("""
                    INSERT INTO bot_admins
                        (telegram_id, role)
                    VALUES
                        (%s, 'admin')
                    ON CONFLICT (telegram_id)
                    DO UPDATE SET role = 'admin'
                """, (admin_id,))

            # Настройка триггеров группы
            cursor.execute("""
                INSERT INTO bot_settings
                    (setting_key, setting_value)
                VALUES
                    ('group_triggers', 'вжух')
                ON CONFLICT (setting_key)
                DO NOTHING
            """)

        db.commit()

        print("DATABASE: подключение успешно")
        print("DATABASE: таблицы готовы")

    except Exception as e:
        db.rollback()
        print("DATABASE INIT ERROR:", repr(e))
        raise

    finally:
        db.close()


# =========================================================
# LOG
# =========================================================

def log_action(admin_id, action, details=""):
    db = None

    try:
        db = get_db()

        with db.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO bot_logs
                    (admin_id, action, details)
                VALUES
                    (%s, %s, %s)
                """,
                (
                    admin_id,
                    action,
                    details
                )
            )

        db.commit()

    except Exception as e:
        print("LOG ERROR:", repr(e))

    finally:
        if db:
            db.close()


# =========================================================
# USERS
# =========================================================

def save_user(user, chat_id=None):
    if not user:
        return

    db = None

    try:
        db = get_db()

        with db.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO bot_users
                    (
                        telegram_id,
                        username,
                        first_name,
                        last_name,
                        chat_id
                    )
                VALUES
                    (%s, %s, %s, %s, %s)
                ON CONFLICT (telegram_id)
                DO UPDATE SET
                    username = EXCLUDED.username,
                    first_name = EXCLUDED.first_name,
                    last_name = EXCLUDED.last_name,
                    chat_id = EXCLUDED.chat_id,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (
                    user.id,
                    user.username,
                    user.first_name,
                    user.last_name,
                    chat_id
                )
            )

        db.commit()

    except Exception as e:
        print("SAVE USER ERROR:", repr(e))

    finally:
        if db:
            db.close()


# =========================================================
# ADMIN CHECK
# =========================================================

def get_admin_role(telegram_id):
    db = None

    try:
        db = get_db()

        with db.cursor() as cursor:
            cursor.execute(
                """
                SELECT role
                FROM bot_admins
                WHERE telegram_id = %s
                """,
                (telegram_id,)
            )

            row = cursor.fetchone()

            if row:
                return row[0]

            return None

    except Exception as e:
        print("ADMIN CHECK ERROR:", repr(e))
        return None

    finally:
        if db:
            db.close()


def is_admin(telegram_id):
    return get_admin_role(telegram_id) in {
        "owner",
        "admin"
    }


def is_owner(telegram_id):
    # Оба заданных администратора имеют одинаковые права.
    return telegram_id in ADMIN_IDS and is_admin(telegram_id)


# =========================================================
# SETTINGS
# =========================================================

def get_setting(key, default=""):
    db = None

    try:
        db = get_db()

        with db.cursor() as cursor:
            cursor.execute(
                """
                SELECT setting_value
                FROM bot_settings
                WHERE setting_key = %s
                """,
                (key,)
            )

            row = cursor.fetchone()

            if row:
                return row[0]

            return default

    except Exception as e:
        print("SETTING ERROR:", repr(e))
        return default

    finally:
        if db:
            db.close()


def set_setting(key, value):
    db = None

    try:
        db = get_db()

        with db.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO bot_settings
                    (setting_key, setting_value)
                VALUES
                    (%s, %s)
                ON CONFLICT (setting_key)
                DO UPDATE SET
                    setting_value = EXCLUDED.setting_value
                """,
                (
                    key,
                    value
                )
            )

        db.commit()

    except Exception as e:
        if db:
            db.rollback()

        print("SETTING SAVE ERROR:", repr(e))

    finally:
        if db:
            db.close()


# =========================================================
# EXCEL
# =========================================================

def is_checked(value):
    if value is None:
        return False

    text = normalize(value)

    checked_values = {
        "да",
        "д",
        "yes",
        "y",
        "1",
        "true",
        "истина",
        "x",
        "х",
        "✓",
        "✔",
        "☑",
        "☒",
        "+",
    }

    return text in checked_values


def import_excel():
    if not os.path.exists(EXCEL_FILE):
        print(
            f"EXCEL: файл {EXCEL_FILE} не найден."
        )
        return

    db = None

    try:
        workbook = load_workbook(
            EXCEL_FILE,
            data_only=True
        )

        sheet = workbook.active

        rows = list(
            sheet.iter_rows(
                values_only=True
            )
        )

        if not rows:
            print("EXCEL: таблица пустая.")
            return

        headers = []

        for value in rows[0]:
            if value is None:
                headers.append("")
            else:
                headers.append(
                    str(value).strip()
                )

        print(
            "EXCEL HEADERS:",
            headers[:20]
        )

        name_index = None
        id_index = None

        for index, header in enumerate(headers):

            normalized_header = normalize(header)

            if normalized_header in {
                "имя",
                "name",
                "игровое имя",
                "ник",
                "никнейм",
            }:
                name_index = index

            if normalized_header in {
                "id",
                "ид",
                "идентификатор",
            }:
                id_index = index

        if name_index is None:
            print(
                "EXCEL ERROR: "
                "не найдена колонка Имя."
            )
            return

        if id_index is None:
            print(
                "EXCEL ERROR: "
                "не найдена колонка ID."
            )
            return

        flower_columns = []

        for index, header in enumerate(headers):

            if index in {
                name_index,
                id_index
            }:
                continue

            if not header:
                continue

            flower_columns.append(
                (index, header)
            )

        print(
            "EXCEL: найдено колонок цветов:",
            len(flower_columns)
        )

        db = get_db()

        added = 0
        already_exists = 0
        skipped = 0

        with db.cursor() as cursor:

            for row in rows[1:]:

                if not row:
                    continue

                person_value = (
                    row[name_index]
                    if name_index < len(row)
                    else None
                )

                flower_id_value = (
                    row[id_index]
                    if id_index < len(row)
                    else None
                )

                if person_value is None:
                    skipped += 1
                    continue

                person = str(
                    person_value
                ).strip()

                if not person:
                    skipped += 1
                    continue

                flower_id = ""

                if flower_id_value is not None:
                    flower_id = str(
                        flower_id_value
                    ).strip()

                # Excel может быть источником старых игроков.
                # Ник не перезаписываем, если ID уже существует в players.
                if flower_id and is_valid_player_id(flower_id):
                    cursor.execute(
                        """
                        INSERT INTO players (player_id, nickname)
                        VALUES (%s, %s)
                        ON CONFLICT (player_id) DO NOTHING
                        """,
                        (flower_id, person)
                    )

                for column_index, flower_name in flower_columns:

                    value = (
                        row[column_index]
                        if column_index < len(row)
                        else None
                    )

                    if not is_checked(value):
                        continue

                    flower = normalize(
                        flower_name
                    )

                    if not flower:
                        continue

                    cursor.execute(
                        """
                        SELECT id
                        FROM flowers
                        WHERE flower = %s
                          AND flower_id = %s
                          AND person = %s
                        LIMIT 1
                        """,
                        (
                            flower,
                            flower_id,
                            person
                        )
                    )

                    if cursor.fetchone():
                        already_exists += 1
                        continue

                    cursor.execute(
                        """
                        INSERT INTO flowers
                            (
                                flower,
                                flower_id,
                                person
                            )
                        VALUES
                            (%s, %s, %s)
                        """,
                        (
                            flower,
                            flower_id,
                            person
                        )
                    )

                    added += 1

        db.commit()

        print("EXCEL: импорт завершён.")
        print("Добавлено:", added)
        print("Уже было:", already_exists)
        print("Пропущено строк:", skipped)

    except Exception as e:

        if db:
            try:
                db.rollback()
            except Exception:
                pass

        print(
            "EXCEL IMPORT ERROR:",
            repr(e)
        )

    finally:

        if db:
            db.close()


# =========================================================
# KEYBOARDS
# =========================================================

def main_keyboard(user_id):
    buttons = [
        [
            InlineKeyboardButton(
                "🆔 Мой ID",
                callback_data="my_id"
            )
        ]
    ]

    if is_admin(user_id):
        buttons.append([
            InlineKeyboardButton(
                "⚙️ Админ-панель",
                callback_data="admin_panel"
            )
        ])

    return InlineKeyboardMarkup(buttons)


def admin_keyboard(user_id):
    buttons = [
        [
            InlineKeyboardButton(
                "👤 Люди",
                callback_data="people_menu"
            ),
            InlineKeyboardButton(
                "🌸 Цветы",
                callback_data="flowers_menu"
            )
        ],
        [
            InlineKeyboardButton(
                "👥 Администраторы",
                callback_data="admins_menu"
            ),
            InlineKeyboardButton(
                "⚡ Группа",
                callback_data="group_menu"
            )
        ],
        [
            InlineKeyboardButton(
                "📊 Статистика",
                callback_data="statistics"
            )
        ],
        [
            InlineKeyboardButton(
                "❌ Закрыть",
                callback_data="close_menu"
            )
        ]
    ]

    return InlineKeyboardMarkup(buttons)


def people_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton(
                "➕ Добавить человека",
                callback_data="person_add"
            )
        ],
        [
            InlineKeyboardButton(
                "✏️ Изменить ник",
                callback_data="person_rename"
            )
        ],
        [
            InlineKeyboardButton(
                "🔎 Найти по ID",
                callback_data="person_find"
            )
        ],
        [
            InlineKeyboardButton(
                "🗑 Удалить человека",
                callback_data="person_delete"
            )
        ],
        [
            InlineKeyboardButton(
                "↩️ Назад",
                callback_data="admin_panel"
            )
        ]
    ])


def flowers_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton(
                "➕ Добавить цветок",
                callback_data="flower_add"
            )
        ],
        [
            InlineKeyboardButton(
                "➖ Удалить цветок",
                callback_data="flower_delete"
            )
        ],
        [
            InlineKeyboardButton(
                "🔎 Найти цветок",
                callback_data="flower_find"
            )
        ],
        [
            InlineKeyboardButton(
                "↩️ Назад",
                callback_data="admin_panel"
            )
        ]
    ])


# =========================================================
# START
# =========================================================

async def start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    user = update.effective_user

    save_user(
        user,
        update.effective_chat.id
    )

    context.user_data.clear()

    await update.message.reply_text(
        "🌸 СКЛАД КУСТИКОВ\n\n"
        "Напиши название цветка, чтобы узнать, "
        "у кого он есть.",
        reply_markup=main_keyboard(user.id)
    )


# =========================================================
# MY ID
# =========================================================

async def show_my_id(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query

    await query.answer()

    await query.message.reply_text(
        f"🆔 Твой Telegram ID:\n\n"
        f"`{query.from_user.id}`",
        parse_mode="Markdown"
    )


# =========================================================
# ADMIN PANEL
# =========================================================

async def show_admin_panel(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query

    await query.answer()

    if not is_admin(query.from_user.id):
        await query.message.reply_text(
            "⛔ Нет доступа."
        )
        return

    context.user_data.clear()

    await query.message.reply_text(
        "⚙️ АДМИН-ПАНЕЛЬ\n\n"
        "Выбери действие:",
        reply_markup=admin_keyboard(
            query.from_user.id
        )
    )


async def admin_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    user = update.effective_user

    save_user(
        user,
        update.effective_chat.id
    )

    if not is_admin(user.id):
        await update.message.reply_text(
            "⛔ Нет доступа."
        )
        return

    context.user_data.clear()

    await update.message.reply_text(
        "⚙️ АДМИН-ПАНЕЛЬ\n\n"
        "Выбери действие:",
        reply_markup=admin_keyboard(user.id)
    )


# =========================================================
# PEOPLE
# =========================================================

async def people_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    await query.message.reply_text(
        "👤 УПРАВЛЕНИЕ ЛЮДЬМИ",
        reply_markup=people_keyboard()
    )


# =========================================================
# ADD PERSON
# =========================================================

async def person_add_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "person_add"
    context.user_data["step"] = "id"

    await query.message.reply_text(
        "➕ Добавление человека\n\n"
        "Напиши ID человека:"
    )


# =========================================================
# RENAME PERSON
# =========================================================

async def person_rename_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "person_rename"
    context.user_data["step"] = "id"

    await query.message.reply_text(
        "✏️ Изменение ника\n\n"
        "Напиши ID человека:"
    )


# =========================================================
# FIND PERSON
# =========================================================

async def person_find_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "person_find"
    context.user_data["step"] = "id"

    await query.message.reply_text(
        "🔎 Напиши ID человека:"
    )


# =========================================================
# DELETE PERSON
# =========================================================

async def person_delete_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "person_delete"
    context.user_data["step"] = "id"

    await query.message.reply_text(
        "🗑 Удаление человека\n\n"
        "Напиши ID человека.\n\n"
        "⚠️ Будут удалены ВСЕ записи этого человека."
    )


# =========================================================
# FLOWER MENU
# =========================================================

async def flowers_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    await query.message.reply_text(
        "🌸 УПРАВЛЕНИЕ ЦВЕТАМИ",
        reply_markup=flowers_keyboard()
    )


# =========================================================
# ADD FLOWER
# =========================================================

async def flower_add_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "flower_add"
    context.user_data["step"] = "flower"

    await query.message.reply_text(
        "➕ Добавление записи\n\n"
        "Напиши название цветка:"
    )


# =========================================================
# DELETE FLOWER
# =========================================================

async def flower_delete_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "flower_delete"
    context.user_data["step"] = "flower"

    await query.message.reply_text(
        "➖ Удаление записи\n\n"
        "Напиши название цветка:"
    )


# =========================================================
# FIND FLOWER
# =========================================================

async def flower_find_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "flower_find"
    context.user_data["step"] = "flower"

    await query.message.reply_text(
        "🔎 Напиши название цветка:"
    )


# =========================================================
# ADMIN MENU
# =========================================================

async def admins_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_owner(query.from_user.id):
        await query.message.reply_text(
            "⛔ Только главный администратор."
        )
        return

    keyboard = [
        [
            InlineKeyboardButton(
                "➕ Добавить админа",
                callback_data="admin_add"
            )
        ],
        [
            InlineKeyboardButton(
                "📋 Список админов",
                callback_data="admin_list"
            )
        ],
        [
            InlineKeyboardButton(
                "🗑 Удалить админа",
                callback_data="admin_delete"
            )
        ],
        [
            InlineKeyboardButton(
                "↩️ Назад",
                callback_data="admin_panel"
            )
        ]
    ]

    await query.message.reply_text(
        "👥 АДМИНИСТРАТОРЫ",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


# =========================================================
# ADD ADMIN
# =========================================================

async def admin_add_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_owner(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "admin_add"
    context.user_data["step"] = "id"

    await query.message.reply_text(
        "➕ Добавление администратора\n\n"
        "Напиши Telegram ID человека.\n\n"
        "Он может получить свой ID кнопкой "
        "«🆔 Мой ID»."
    )


# =========================================================
# DELETE ADMIN
# =========================================================

async def admin_delete_start(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_owner(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "admin_delete"
    context.user_data["step"] = "id"

    await query.message.reply_text(
        "🗑 Удаление администратора\n\n"
        "Напиши Telegram ID."
    )


# =========================================================
# ADMIN LIST
# =========================================================

async def admin_list(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_owner(query.from_user.id):
        return

    db = None

    try:
        db = get_db()

        with db.cursor() as cursor:
            cursor.execute(
                """
                SELECT telegram_id, role
                FROM bot_admins
                ORDER BY
                    CASE
                        WHEN role = 'owner'
                        THEN 0
                        ELSE 1
                    END,
                    telegram_id
                """
            )

            rows = cursor.fetchall()

    except Exception as e:
        print("ADMIN LIST ERROR:", repr(e))
        await query.message.reply_text(
            "❌ Ошибка базы данных."
        )
        return

    finally:
        if db:
            db.close()

    if not rows:
        await query.message.reply_text(
            "Администраторов нет."
        )
        return

    result = "👥 АДМИНИСТРАТОРЫ\n\n"

    for telegram_id, role in rows:

        if role == "owner":
            title = "👑 Главный админ"
        else:
            title = "🛠 Администратор"

        result += (
            f"{title}\n"
            f"🆔 {telegram_id}\n\n"
        )

    await query.message.reply_text(result)


# =========================================================
# GROUP SETTINGS
# =========================================================

async def group_menu(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    triggers = get_setting(
        "group_triggers",
        "вжух"
    )

    keyboard = [
        [
            InlineKeyboardButton(
                "⚡ Изменить слова",
                callback_data="group_triggers_edit"
            )
        ],
        [
            InlineKeyboardButton(
                "↩️ Назад",
                callback_data="admin_panel"
            )
        ]
    ]

    await query.message.reply_text(
        "⚡ НАСТРОЙКИ ГРУППЫ\n\n"
        "Сейчас бот реагирует на:\n"
        f"• {triggers}\n\n"
        "В группе бот будет молчать на обычные "
        "сообщения и отвечать только на сообщение "
        "с одним из этих слов + названием цветка.",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def group_triggers_edit(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    context.user_data.clear()
    context.user_data["action"] = "group_triggers"
    context.user_data["step"] = "triggers"

    await query.message.reply_text(
        "⚡ Напиши слова-триггеры через запятую.\n\n"
        "Например:\n"
        "вжух, нужна, ищу\n\n"
        "Тогда бот будет реагировать на:\n"
        "«вжух роза»\n"
        "«нужна роза»\n"
        "«ищу розу»"
    )


# =========================================================
# STATISTICS
# =========================================================

async def statistics(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    db = None

    try:
        db = get_db()

        with db.cursor() as cursor:

            cursor.execute(
                "SELECT COUNT(*) FROM flowers"
            )
            total_records = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT COUNT(DISTINCT flower_id)
                FROM flowers
                WHERE flower_id IS NOT NULL
                  AND flower_id <> ''
                """
            )
            people = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT COUNT(DISTINCT flower)
                FROM flowers
                """
            )
            flowers = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT COUNT(*)
                FROM bot_admins
                """
            )
            admins = cursor.fetchone()[0]

    except Exception as e:
        print("STATISTICS ERROR:", repr(e))
        await query.message.reply_text(
            "❌ Ошибка базы данных."
        )
        return

    finally:
        if db:
            db.close()

    await query.message.reply_text(
        "📊 СТАТИСТИКА\n\n"
        f"🌸 Записей о цветах: {total_records}\n"
        f"👤 Людей по ID: {people}\n"
        f"🌺 Разных цветов: {flowers}\n"
        f"🛠 Администраторов: {admins}"
    )


# =========================================================
# CONFIRM DELETE PERSON
# =========================================================

def delete_person_confirm_keyboard():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton(
                "🗑 ДА, УДАЛИТЬ",
                callback_data="confirm_person_delete"
            )
        ],
        [
            InlineKeyboardButton(
                "❌ Отмена",
                callback_data="cancel_action"
            )
        ]
    ])


# =========================================================
# ADMIN TEXT INPUT
# =========================================================

async def admin_input(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    user = update.effective_user
    chat_id = update.effective_chat.id

    if not is_admin(user.id):
        return

    text = update.message.text.strip()

    if not text:
        return

    # =====================================================
    # /at и /add: отдельное состояние по user_id + chat_id
    # =====================================================
    command_state = get_command_state(update)

    if command_state:
        command = command_state.get("command")
        step = command_state.get("step")

        if command == "at":
            if step == "id":
                if not is_valid_player_id(text):
                    await update.message.reply_text(
                        "❌ ID должен состоять только из латинских букв и цифр.\n\n"
                        "Пример: s18 или ee15z72gm9u"
                    )
                    return

                db = None
                try:
                    db = get_db()
                    with db.cursor() as cursor:
                        cursor.execute(
                            "SELECT nickname FROM players WHERE player_id = %s",
                            (text,)
                        )
                        row = cursor.fetchone()

                        if row:
                            await update.message.reply_text(
                                f"⚠️ Игрок с ID {text} уже существует.\n"
                                f"👤 Ник: {row[0]}"
                            )
                            clear_command_state(update)
                            return

                    command_state["player_id"] = text
                    command_state["step"] = "nickname"

                    await update.message.reply_text(
                        "Введите ник игрока:"
                    )
                except Exception as e:
                    print("AT ID ERROR:", repr(e))
                    await update.message.reply_text("❌ Ошибка базы данных.")
                finally:
                    if db:
                        db.close()
                return

            if step == "nickname":
                player_id = command_state.get("player_id")
                nickname = text

                if not player_id:
                    await update.message.reply_text("❌ Операция устарела. Запусти /at заново.")
                    clear_command_state(update)
                    return

                db = None
                try:
                    db = get_db()
                    with db.cursor() as cursor:
                        cursor.execute(
                            """
                            INSERT INTO players (player_id, nickname)
                            VALUES (%s, %s)
                            ON CONFLICT (player_id) DO NOTHING
                            """,
                            (player_id, nickname)
                        )
                        created = cursor.rowcount
                    db.commit()

                    if not created:
                        await update.message.reply_text(
                            f"⚠️ Игрок с ID {player_id} уже существует."
                        )
                    else:
                        log_action(
                            user.id,
                            "ADD_PLAYER_COMMAND",
                            f"{player_id} | {nickname}"
                        )
                        await update.message.reply_text(
                            f"✅ Игрок {player_id} — {nickname} успешно добавлен."
                        )
                except Exception as e:
                    if db:
                        db.rollback()
                    print("AT ADD ERROR:", repr(e))
                    await update.message.reply_text("❌ Ошибка базы данных.")
                finally:
                    if db:
                        db.close()

                clear_command_state(update)
                return

        if command == "delete":
            if step == "id":
                if not is_valid_player_id(text):
                    await update.message.reply_text(
                        "❌ ID должен состоять только из латинских букв и цифр."
                    )
                    return

                player_id = text
                db = None
                try:
                    db = get_db()
                    with db.cursor() as cursor:
                        cursor.execute(
                            "SELECT nickname FROM players WHERE player_id = %s",
                            (player_id,)
                        )
                        row = cursor.fetchone()
                        if not row:
                            await update.message.reply_text(
                                f"❌ ID {player_id} не найден."
                            )
                            clear_command_state(update)
                            return

                        cursor.execute(
                            "SELECT COUNT(*) FROM flowers WHERE flower_id = %s",
                            (player_id,)
                        )
                        flower_count = cursor.fetchone()[0]

                    command_state["player_id"] = player_id
                    command_state["nickname"] = row[0]
                    command_state["flower_count"] = flower_count
                    command_state["step"] = "confirm"

                    await update.message.reply_text(
                        "⚠️ ВНИМАНИЕ\n\n"
                        f"🆔 ID: {player_id}\n"
                        f"👤 {row[0]}\n"
                        f"🌸 Цветов: {flower_count}\n\n"
                        "Для подтверждения напиши: ДА\n"
                        "Для отмены: НЕТ"
                    )
                except Exception as e:
                    print("DELETE COMMAND CHECK ERROR:", repr(e))
                    await update.message.reply_text("❌ Ошибка базы данных.")
                finally:
                    if db:
                        db.close()
                return

            if step == "confirm":
                answer = normalize(text)
                if answer not in {"да", "нет", "д", "н"}:
                    await update.message.reply_text("Введите ДА для удаления или НЕТ для отмены.")
                    return

                if answer in {"нет", "н"}:
                    clear_command_state(update)
                    await update.message.reply_text("❌ Удаление отменено.")
                    return

                player_id = command_state.get("player_id")
                db = None
                try:
                    db = get_db()
                    with db.cursor() as cursor:
                        cursor.execute("DELETE FROM flowers WHERE flower_id = %s", (player_id,))
                        deleted_flowers = cursor.rowcount
                        cursor.execute("DELETE FROM players WHERE player_id = %s", (player_id,))
                        deleted_player = cursor.rowcount
                    db.commit()

                    log_action(
                        user.id,
                        "DELETE_PERSON_COMMAND",
                        f"ID {player_id}, player {deleted_player}, records {deleted_flowers}"
                    )

                    await update.message.reply_text(
                        "✅ Человек полностью удалён.\n\n"
                        f"🆔 ID: {player_id}\n"
                        f"🌸 Удалено записей о цветах: {deleted_flowers}"
                    )
                except Exception as e:
                    if db:
                        db.rollback()
                    print("DELETE COMMAND ERROR:", repr(e))
                    await update.message.reply_text("❌ Ошибка базы данных.")
                finally:
                    if db:
                        db.close()

                clear_command_state(update)
                return

        if command == "add":
            if step == "flower":
                flower = normalize(text)
                if not flower:
                    await update.message.reply_text("❌ Название цветка не может быть пустым.")
                    return
                command_state["flower"] = flower
                command_state["step"] = "id"
                await update.message.reply_text("Введите ID игрока:")
                return

            if step == "id":
                if not is_valid_player_id(text):
                    await update.message.reply_text(
                        "❌ ID должен состоять только из латинских букв и цифр."
                    )
                    return

                flower = command_state.get("flower")
                db = None
                try:
                    db = get_db()
                    with db.cursor() as cursor:
                        cursor.execute(
                            "SELECT nickname FROM players WHERE player_id = %s",
                            (text,)
                        )
                        row = cursor.fetchone()

                        if not row:
                            await update.message.reply_text(
                                f"❌ Игрок с ID {text} не найден.\n\n"
                                "Сначала добавь игрока командой /at."
                            )
                            clear_command_state(update)
                            return

                        nickname = row[0]

                        cursor.execute(
                            """
                            SELECT id
                            FROM flowers
                            WHERE flower = %s
                              AND flower_id = %s
                            LIMIT 1
                            """,
                            (flower, text)
                        )
                        if cursor.fetchone():
                            await update.message.reply_text(
                                "⚠️ Такой цветок уже добавлен этому игроку."
                            )
                            clear_command_state(update)
                            return

                        cursor.execute(
                            """
                            INSERT INTO flowers (flower, flower_id, person)
                            VALUES (%s, %s, %s)
                            """,
                            (flower, text, nickname)
                        )

                    db.commit()

                    log_action(
                        user.id,
                        "ADD_FLOWER_COMMAND",
                        f"{flower} | {text} | {nickname}"
                    )

                    await update.message.reply_text(
                        "✅ Цветок успешно добавлен.\n\n"
                        f"🌸 {flower}\n"
                        f"🆔 {text}\n"
                        f"👤 {nickname}"
                    )
                except Exception as e:
                    if db:
                        db.rollback()
                    print("ADD COMMAND ERROR:", repr(e))
                    await update.message.reply_text("❌ Ошибка базы данных.")
                finally:
                    if db:
                        db.close()

                clear_command_state(update)
                return

        return

    # -----------------------------------------------------
    # EXISTING ADMIN OPERATIONS FROM BUTTONS
    # -----------------------------------------------------

    action = context.user_data.get("action")
    step = context.user_data.get("step")

    if not action or not step:
        return

    # -----------------------------------------------------
    # ADD ADMIN
    # -----------------------------------------------------
    if action == "admin_add":
        if not is_owner(user.id):
            context.user_data.clear()
            return
        if step == "id":
            if not text.isdigit():
                await update.message.reply_text("❌ ID должен состоять только из цифр.")
                return
            telegram_id = int(text)
            if is_owner(telegram_id):
                await update.message.reply_text("👑 Это уже главный администратор.")
                context.user_data.clear()
                return
            db = None
            try:
                db = get_db()
                with db.cursor() as cursor:
                    cursor.execute(
                        """
                        INSERT INTO bot_admins (telegram_id, role, added_by)
                        VALUES (%s, 'admin', %s)
                        ON CONFLICT (telegram_id) DO UPDATE SET role = 'admin'
                        """,
                        (telegram_id, user.id)
                    )
                db.commit()
                log_action(user.id, "ADD_ADMIN", str(telegram_id))
                await update.message.reply_text(
                    "✅ Администратор добавлен.\n\n"
                    f"🆔 {telegram_id}"
                )
            except Exception as e:
                if db:
                    db.rollback()
                print("ADD ADMIN ERROR:", repr(e))
                await update.message.reply_text("❌ Ошибка базы данных.")
            finally:
                if db:
                    db.close()
            context.user_data.clear()
            return

    # -----------------------------------------------------
    # DELETE ADMIN
    # -----------------------------------------------------
    if action == "admin_delete":
        if not is_owner(user.id):
            context.user_data.clear()
            return
        if step == "id":
            if not text.isdigit():
                await update.message.reply_text("❌ ID должен состоять только из цифр.")
                return
            telegram_id = int(text)
            if telegram_id in ADMIN_IDS:
                await update.message.reply_text("⛔ Постоянного администратора удалить нельзя.")
                context.user_data.clear()
                return
            db = None
            try:
                db = get_db()
                with db.cursor() as cursor:
                    cursor.execute(
                        "DELETE FROM bot_admins WHERE telegram_id = %s AND role <> 'owner'",
                        (telegram_id,)
                    )
                    deleted = cursor.rowcount
                db.commit()
                if deleted:
                    log_action(user.id, "DELETE_ADMIN", str(telegram_id))
                    await update.message.reply_text("✅ Администратор удалён.")
                else:
                    await update.message.reply_text("❌ Такой администратор не найден.")
            except Exception as e:
                if db:
                    db.rollback()
                print("DELETE ADMIN ERROR:", repr(e))
                await update.message.reply_text("❌ Ошибка базы данных.")
            finally:
                if db:
                    db.close()
            context.user_data.clear()
            return

    # -----------------------------------------------------
    # GROUP TRIGGERS
    # -----------------------------------------------------
    if action == "group_triggers":
        if step == "triggers":
            parts = [normalize(x) for x in text.split(",") if normalize(x)]
            if not parts:
                await update.message.reply_text("❌ Нужно указать хотя бы одно слово.")
                return
            parts = list(dict.fromkeys(parts))
            value = ", ".join(parts)
            set_setting("group_triggers", value)
            log_action(user.id, "CHANGE_TRIGGERS", value)
            await update.message.reply_text(
                "✅ Триггеры сохранены.\n\n" + "\n".join(f"⚡ {x}" for x in parts)
            )
            context.user_data.clear()
            return

    # -----------------------------------------------------
    # FIND PERSON
    # -----------------------------------------------------
    if action == "person_find" and step == "id":
        if not is_valid_player_id(text):
            await update.message.reply_text("❌ ID должен состоять только из латинских букв и цифр.")
            return
        player_id = text
        db = None
        try:
            db = get_db()
            with db.cursor() as cursor:
                cursor.execute(
                    "SELECT nickname FROM players WHERE player_id = %s",
                    (player_id,)
                )
                player = cursor.fetchone()
                cursor.execute(
                    """
                    SELECT flower
                    FROM flowers
                    WHERE flower_id = %s
                    ORDER BY flower
                    """,
                    (player_id,)
                )
                flowers = cursor.fetchall()
        except Exception as e:
            print("PERSON FIND ERROR:", repr(e))
            await update.message.reply_text("❌ Ошибка базы данных.")
            context.user_data.clear()
            return
        finally:
            if db:
                db.close()

        if not player:
            await update.message.reply_text(f"❌ ID {player_id} не найден.")
            context.user_data.clear()
            return

        result = f"🆔 ID: {player_id}\n\n👤 {player[0]}\n\n🌸 Цветы:\n"
        if flowers:
            result += "\n".join(f"• {row[0]}" for row in flowers)
        else:
            result += "• Цветов нет"
        await update.message.reply_text(result)
        context.user_data.clear()
        return

    # -----------------------------------------------------
    # DELETE PERSON
    # -----------------------------------------------------
    if action == "person_delete" and step == "id":
        if not is_valid_player_id(text):
            await update.message.reply_text("❌ ID должен состоять только из латинских букв и цифр.")
            return
        player_id = text
        db = None
        try:
            db = get_db()
            with db.cursor() as cursor:
                cursor.execute(
                    "SELECT nickname FROM players WHERE player_id = %s",
                    (player_id,)
                )
                player = cursor.fetchone()
                if not player:
                    await update.message.reply_text(f"❌ ID {player_id} не найден.")
                    context.user_data.clear()
                    return
                cursor.execute(
                    "SELECT COUNT(*) FROM flowers WHERE flower_id = %s",
                    (player_id,)
                )
                flower_count = cursor.fetchone()[0]
        except Exception as e:
            print("PERSON DELETE CHECK ERROR:", repr(e))
            await update.message.reply_text("❌ Ошибка базы данных.")
            context.user_data.clear()
            return
        finally:
            if db:
                db.close()

        context.user_data["delete_id"] = player_id
        await update.message.reply_text(
            "⚠️ ВНИМАНИЕ\n\n"
            f"ID: {player_id}\n"
            f"👤 {player[0]}\n"
            f"🌸 Цветов: {flower_count}\n\n"
            "Будут удалены игрок и ВСЕ его записи о цветах.\n"
            "Это действие нельзя отменить.",
            reply_markup=delete_person_confirm_keyboard()
        )
        return

    # -----------------------------------------------------
    # RENAME PERSON
    # -----------------------------------------------------
    if action == "person_rename":
        if step == "id":
            if not is_valid_player_id(text):
                await update.message.reply_text("❌ ID должен состоять только из латинских букв и цифр.")
                return
            db = None
            try:
                db = get_db()
                with db.cursor() as cursor:
                    cursor.execute("SELECT nickname FROM players WHERE player_id = %s", (text,))
                    row = cursor.fetchone()
                if not row:
                    await update.message.reply_text("❌ Такой ID не найден.")
                    context.user_data.clear()
                    return
            except Exception as e:
                print("RENAME CHECK ERROR:", repr(e))
                await update.message.reply_text("❌ Ошибка базы данных.")
                context.user_data.clear()
                return
            finally:
                if db:
                    db.close()
            context.user_data["rename_id"] = text
            context.user_data["step"] = "name"
            await update.message.reply_text("✏️ Теперь напиши новое игровое имя:")
            return

        if step == "name":
            player_id = context.user_data["rename_id"]
            new_name = text
            db = None
            try:
                db = get_db()
                with db.cursor() as cursor:
                    cursor.execute("SELECT nickname FROM players WHERE player_id = %s", (player_id,))
                    row = cursor.fetchone()
                    if not row:
                        await update.message.reply_text("❌ Такой ID не найден.")
                        context.user_data.clear()
                        return
                    old_name = row[0]
                    cursor.execute(
                        """
                        UPDATE players
                        SET nickname = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE player_id = %s
                        """,
                        (new_name, player_id)
                    )
                    # Сохраняем старое поле person совместимым со старой системой.
                    cursor.execute(
                        "UPDATE flowers SET person = %s WHERE flower_id = %s",
                        (new_name, player_id)
                    )
                    changed = cursor.rowcount
                db.commit()
                log_action(user.id, "RENAME_PERSON", f"{player_id}: {old_name} -> {new_name}")
                await update.message.reply_text(
                    "✅ Имя изменено.\n\n"
                    f"🆔 ID: {player_id}\n"
                    f"Было: {old_name}\n"
                    f"Стало: {new_name}\n\n"
                    f"Цветочных записей обновлено: {changed}"
                )
            except Exception as e:
                if db:
                    db.rollback()
                print("RENAME ERROR:", repr(e))
                await update.message.reply_text("❌ Ошибка базы данных.")
            finally:
                if db:
                    db.close()
            context.user_data.clear()
            return

    # -----------------------------------------------------
    # ADD PERSON FROM EXISTING BUTTON
    # -----------------------------------------------------
    if action == "person_add":
        if step == "id":
            if not is_valid_player_id(text):
                await update.message.reply_text("❌ ID должен состоять только из латинских букв и цифр.")
                return
            db = None
            try:
                db = get_db()
                with db.cursor() as cursor:
                    cursor.execute("SELECT nickname FROM players WHERE player_id = %s", (text,))
                    if cursor.fetchone():
                        await update.message.reply_text("⚠️ Игрок с таким ID уже существует.")
                        context.user_data.clear()
                        return
            except Exception as e:
                print("ADD PERSON CHECK ERROR:", repr(e))
                await update.message.reply_text("❌ Ошибка базы данных.")
                context.user_data.clear()
                return
            finally:
                if db:
                    db.close()
            context.user_data["new_id"] = text
            context.user_data["step"] = "name"
            await update.message.reply_text("Теперь напиши игровое имя:")
            return

        if step == "name":
            context.user_data["new_person"] = text
            context.user_data["step"] = "flower"
            await update.message.reply_text("Теперь напиши название цветка:")
            return

        if step == "flower":
            flower = normalize(text)
            player_id = context.user_data["new_id"]
            person = context.user_data["new_person"]
            db = None
            try:
                db = get_db()
                with db.cursor() as cursor:
                    cursor.execute(
                        "INSERT INTO players (player_id, nickname) VALUES (%s, %s)",
                        (player_id, person)
                    )
                    cursor.execute(
                        """
                        INSERT INTO flowers (flower, flower_id, person)
                        VALUES (%s, %s, %s)
                        """,
                        (flower, player_id, person)
                    )
                db.commit()
                log_action(user.id, "ADD_PERSON_FLOWER", f"{player_id} | {person} | {flower}")
                await update.message.reply_text(
                    "✅ Добавлено:\n\n"
                    f"🆔 {player_id}\n👤 {person}\n🌸 {flower}"
                )
            except Exception as e:
                if db:
                    db.rollback()
                print("ADD PERSON ERROR:", repr(e))
                await update.message.reply_text("❌ Ошибка базы данных.")
            finally:
                if db:
                    db.close()
            context.user_data.clear()
            return

    # -----------------------------------------------------
    # ADD FLOWER FROM EXISTING BUTTON
    # -----------------------------------------------------
    if action == "flower_add":
        if step == "flower":
            flower = normalize(text)
            if not flower:
                await update.message.reply_text("❌ Название цветка не может быть пустым.")
                return
            context.user_data["flower"] = flower
            context.user_data["step"] = "id"
            await update.message.reply_text("Теперь напиши ID игрока:")
            return

        if step == "id":
            if not is_valid_player_id(text):
                await update.message.reply_text("❌ ID должен состоять только из латинских букв и цифр.")
                return
            flower = context.user_data["flower"]
            db = None
            try:
                db = get_db()
                with db.cursor() as cursor:
                    cursor.execute("SELECT nickname FROM players WHERE player_id = %s", (text,))
                    row = cursor.fetchone()
                    if not row:
                        await update.message.reply_text(
                            f"❌ Игрок с ID {text} не найден.\n\nСначала добавь игрока через /at."
                        )
                        context.user_data.clear()
                        return
                    nickname = row[0]
                    cursor.execute(
                        "SELECT id FROM flowers WHERE flower = %s AND flower_id = %s LIMIT 1",
                        (flower, text)
                    )
                    if cursor.fetchone():
                        await update.message.reply_text("⚠️ Такой цветок уже добавлен этому игроку.")
                        context.user_data.clear()
                        return
                    cursor.execute(
                        "INSERT INTO flowers (flower, flower_id, person) VALUES (%s, %s, %s)",
                        (flower, text, nickname)
                    )
                db.commit()
                log_action(user.id, "ADD_FLOWER", f"{flower} | {text} | {nickname}")
                await update.message.reply_text(
                    "✅ Добавлено:\n\n"
                    f"🌸 {flower}\n🆔 {text}\n👤 {nickname}"
                )
            except Exception as e:
                if db:
                    db.rollback()
                print("ADD FLOWER ERROR:", repr(e))
                await update.message.reply_text("❌ Ошибка базы данных.")
            finally:
                if db:
                    db.close()
            context.user_data.clear()
            return

    # -----------------------------------------------------
    # DELETE FLOWER FROM EXISTING BUTTON
    # -----------------------------------------------------
    if action == "flower_delete":
        if step == "flower":
            context.user_data["flower"] = normalize(text)
            context.user_data["step"] = "id"
            await update.message.reply_text("Теперь напиши ID игрока:")
            return

        if step == "id":
            if not is_valid_player_id(text):
                await update.message.reply_text("❌ ID должен состоять только из латинских букв и цифр.")
                return
            flower = context.user_data["flower"]
            db = None
            try:
                db = get_db()
                with db.cursor() as cursor:
                    cursor.execute(
                        "DELETE FROM flowers WHERE flower = %s AND flower_id = %s",
                        (flower, text)
                    )
                    deleted = cursor.rowcount
                    cursor.execute("SELECT nickname FROM players WHERE player_id = %s", (text,))
                    row = cursor.fetchone()
                    nickname = row[0] if row else ""
                db.commit()
                if deleted:
                    log_action(user.id, "DELETE_FLOWER", f"{flower} | {text} | {nickname}")
                    await update.message.reply_text(
                        "✅ Удалено:\n\n"
                        f"🌸 {flower}\n🆔 {text}\n👤 {nickname}"
                    )
                else:
                    await update.message.reply_text("❌ Такая запись не найдена.")
            except Exception as e:
                if db:
                    db.rollback()
                print("DELETE FLOWER ERROR:", repr(e))
                await update.message.reply_text("❌ Ошибка базы данных.")
            finally:
                if db:
                    db.close()
            context.user_data.clear()
            return

    # -----------------------------------------------------
    # FIND FLOWER
    # -----------------------------------------------------
    if action == "flower_find" and step == "flower":
        flower = normalize(text)
        db = None
        try:
            db = get_db()
            with db.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT f.flower_id, COALESCE(p.nickname, f.person)
                    FROM flowers f
                    LEFT JOIN players p ON p.player_id = f.flower_id
                    WHERE f.flower = %s
                    ORDER BY COALESCE(p.nickname, f.person)
                    """,
                    (flower,)
                )
                rows = cursor.fetchall()
        except Exception as e:
            print("FLOWER FIND ERROR:", repr(e))
            await update.message.reply_text("❌ Ошибка базы данных.")
            context.user_data.clear()
            return
        finally:
            if db:
                db.close()
        if not rows:
            await update.message.reply_text(f"❌ Не нашёл «{text}».")
            context.user_data.clear()
            return
        result = f"🌸 {text}\n\nЕсть у:\n"
        for flower_id, person in rows:
            result += f"🆔 {flower_id or 'ID не указан'} — 👤 {person}\n"
        await update.message.reply_text(result)
        context.user_data.clear()
        return


# =========================================================
# CONFIRM DELETE PERSON
# =========================================================

async def confirm_person_delete(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query
    await query.answer()

    if not is_admin(query.from_user.id):
        return

    player_id = context.user_data.get("delete_id")

    if not player_id:
        await query.message.reply_text("❌ Операция устарела.")
        context.user_data.clear()
        return

    db = None
    try:
        db = get_db()
        with db.cursor() as cursor:
            cursor.execute("DELETE FROM flowers WHERE flower_id = %s", (player_id,))
            deleted_flowers = cursor.rowcount
            cursor.execute("DELETE FROM players WHERE player_id = %s", (player_id,))
            deleted_player = cursor.rowcount
        db.commit()

        log_action(
            query.from_user.id,
            "DELETE_PERSON",
            f"ID {player_id}, player {deleted_player}, records {deleted_flowers}"
        )

        await query.message.reply_text(
            "✅ Человек полностью удалён.\n\n"
            f"🆔 ID: {player_id}\n"
            f"🌸 Удалено записей о цветах: {deleted_flowers}"
        )
    except Exception as e:
        if db:
            db.rollback()
        print("DELETE PERSON ERROR:", repr(e))
        await query.message.reply_text("❌ Ошибка базы данных.")
    finally:
        if db:
            db.close()

    context.user_data.clear()


# =========================================================
# CANCEL
# =========================================================

async def cancel_action(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query

    await query.answer(
        "Отменено"
    )

    context.user_data.clear()
    clear_command_state(update)

    await query.message.reply_text(
        "❌ Операция отменена."
    )


# =========================================================
# GENERAL CALLBACKS
# =========================================================

async def callback_handler(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    query = update.callback_query

    data = query.data

    if data == "my_id":
        await show_my_id(
            update,
            context
        )
        return

    if data == "admin_panel":
        await show_admin_panel(
            update,
            context
        )
        return

    if data == "people_menu":
        await people_menu(
            update,
            context
        )
        return

    if data == "person_add":
        await person_add_start(
            update,
            context
        )
        return

    if data == "person_rename":
        await person_rename_start(
            update,
            context
        )
        return

    if data == "person_find":
        await person_find_start(
            update,
            context
        )
        return

    if data == "person_delete":
        await person_delete_start(
            update,
            context
        )
        return

    if data == "flowers_menu":
        await flowers_menu(
            update,
            context
        )
        return

    if data == "flower_add":
        await flower_add_start(
            update,
            context
        )
        return

    if data == "flower_delete":
        await flower_delete_start(
            update,
            context
        )
        return

    if data == "flower_find":
        await flower_find_start(
            update,
            context
        )
        return

    if data == "admins_menu":
        await admins_menu(
            update,
            context
        )
        return

    if data == "admin_add":
        await admin_add_start(
            update,
            context
        )
        return

    if data == "admin_delete":
        await admin_delete_start(
            update,
            context
        )
        return

    if data == "admin_list":
        await admin_list(
            update,
            context
        )
        return

    if data == "group_menu":
        await group_menu(
            update,
            context
        )
        return

    if data == "group_triggers_edit":
        await group_triggers_edit(
            update,
            context
        )
        return

    if data == "statistics":
        await statistics(
            update,
            context
        )
        return

    if data == "confirm_person_delete":
        await confirm_person_delete(
            update,
            context
        )
        return

    if data == "cancel_action":
        await cancel_action(
            update,
            context
        )
        return

    if data == "close_menu":

        await query.answer(
            "Меню закрыто"
        )

        context.user_data.clear()

        try:
            await query.message.delete()
        except Exception:
            pass

        return

    await query.answer()


# =========================================================
# NORMAL PRIVATE SEARCH
# =========================================================

async def search_flower(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    user = update.effective_user

    if not user:
        return

    # В группе обычные сообщения полностью игнорируем.
    if update.effective_chat.type in {
        "group",
        "supergroup"
    }:
        return

    if is_admin(user.id):
        if context.user_data.get("action"):
            return

    text = update.message.text.strip()

    if not text or text.startswith("/"):
        return

    flower = normalize(text)

    db = None

    try:

        db = get_db()

        with db.cursor() as cursor:

            cursor.execute(
                """
                SELECT f.flower_id, COALESCE(p.nickname, f.person)
                FROM flowers f
                LEFT JOIN players p ON p.player_id = f.flower_id
                WHERE f.flower = %s
                ORDER BY COALESCE(p.nickname, f.person)
                """,
                (flower,)
            )

            rows = cursor.fetchall()

    except Exception as e:

        print(
            "DATABASE SEARCH ERROR:",
            repr(e)
        )

        await update.message.reply_text(
            "❌ Ошибка при обращении к базе данных."
        )

        return

    finally:

        if db:
            db.close()

    if not rows:

        await update.message.reply_text(
            f"❌ Не нашёл «{text}»."
        )

        return

    result = (
        f"🌸 {text}\n\n"
        "Есть у:\n"
    )

    for flower_id, person in rows:

        if flower_id:

            result += (
                f"🆔 {flower_id} — "
                f"👤 {person}\n"
            )

        else:

            result += (
                f"🆔 ID не указан — "
                f"👤 {person}\n"
            )

    await update.message.reply_text(
        result
    )


# =========================================================
# GROUP TRIGGER
# =========================================================

async def group_trigger(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    if not update.message:
        return

    if update.effective_chat.type not in {
        "group",
        "supergroup"
    }:
        return

    text = update.message.text

    if not text:
        return

    original = text.strip()
    normalized = normalize(original)

    triggers_raw = get_setting(
        "group_triggers",
        "вжух"
    )

    triggers = [
        normalize(x)
        for x in triggers_raw.split(",")
        if normalize(x)
    ]

    matched_trigger = None
    flower_query = ""

    # Поддерживаем:
    # вжух роза
    # вжух, роза
    # вжух: роза
    # вжух — роза

    for trigger in triggers:

        pattern = (
            r"^"
            + re.escape(trigger)
            + r"(?:\s*[,:\-—]\s*|\s+)"
            + r"(.+)$"
        )

        match = re.match(
            pattern,
            normalized,
            flags=re.IGNORECASE
        )

        if match:

            matched_trigger = trigger
            flower_query = match.group(1).strip()

            break

    if not matched_trigger:
        return

    if not flower_query:
        return

    db = None

    try:

        db = get_db()

        with db.cursor() as cursor:

            cursor.execute(
                """
                SELECT f.flower_id, COALESCE(p.nickname, f.person)
                FROM flowers f
                LEFT JOIN players p ON p.player_id = f.flower_id
                WHERE f.flower = %s
                ORDER BY COALESCE(p.nickname, f.person)
                """,
                (flower_query,)
            )

            rows = cursor.fetchall()

    except Exception as e:

        print(
            "GROUP SEARCH ERROR:",
            repr(e)
        )

        return

    finally:

        if db:
            db.close()

    if not rows:

        await update.message.reply_text(
            f"❌ Не нашёл «{flower_query}»."
        )

        return

    result = (
        f"🌸 {flower_query}\n\n"
        "Есть у:\n"
    )

    for flower_id, person in rows:

        if flower_id:

            result += (
                f"🆔 {flower_id} — "
                f"👤 {person}\n"
            )

        else:

            result += (
                f"🆔 ID не указан — "
                f"👤 {person}\n"
            )

    await update.message.reply_text(
        result
    )


# =========================================================
# /DELETE — DELETE PLAYER BY PERMANENT ID
# =========================================================

async def delete_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    user = update.effective_user

    if not is_admin(user.id):
        await update.message.reply_text("⛔ Нет доступа.")
        return

    COMMAND_STATES[command_state_key(update)] = {
        "command": "delete",
        "step": "id"
    }

    await update.message.reply_text(
        "🗑 Удаление человека\n\n"
        "Введите ID игрока.\n\n"
        "⚠️ Будут удалены игрок и ВСЕ его записи о цветах."
    )


# =========================================================
# /AT — CREATE PLAYER
# =========================================================

async def at_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    user = update.effective_user

    if not is_admin(user.id):
        await update.message.reply_text("⛔ Нет доступа.")
        return

    # Новая команда всегда начинает собственный диалог.
    COMMAND_STATES[command_state_key(update)] = {
        "command": "at",
        "step": "id"
    }

    await update.message.reply_text("Введите ID игрока:")


# =========================================================
# /ADD — ADD FLOWER TO EXISTING PLAYER
# =========================================================

async def add_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    user = update.effective_user

    if not is_admin(user.id):
        await update.message.reply_text("⛔ Нет доступа.")
        return

    key = command_state_key(update)
    args = context.args or []

    if args:
        flower = normalize(" ".join(args))
        COMMAND_STATES[key] = {
            "command": "add",
            "step": "id",
            "flower": flower
        }
        await update.message.reply_text(
            f"🌸 Цветок: {flower}\n\nВведите ID игрока:"
        )
    else:
        COMMAND_STATES[key] = {
            "command": "add",
            "step": "flower"
        }
        await update.message.reply_text("Введите название цветка:")


# =========================================================
# CANCEL COMMAND
# =========================================================

async def cancel_command(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE
):
    context.user_data.clear()
    clear_command_state(update)

    await update.message.reply_text(
        "❌ Операция отменена."
    )


# =========================================================
# TELEGRAM HANDLERS
# =========================================================

telegram_app.add_handler(
    CommandHandler(
        "start",
        start
    )
)

telegram_app.add_handler(
    CommandHandler(
        "admin",
        admin_command
    )
)

telegram_app.add_handler(
    CommandHandler(
        "delete",
        delete_command
    )
)

telegram_app.add_handler(
    CommandHandler(
        "at",
        at_command
    )
)

telegram_app.add_handler(
    CommandHandler(
        "add",
        add_command
    )
)

telegram_app.add_handler(
    CommandHandler(
        "cancel",
        cancel_command
    )
)

telegram_app.add_handler(
    CallbackQueryHandler(
        callback_handler
    )
)

# Группа 0:
# админские операции
telegram_app.add_handler(
    MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        admin_input
    ),
    group=0
)

# Группа 1:
# реакция в группе
telegram_app.add_handler(
    MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        group_trigger
    ),
    group=1
)

# Группа 2:
# обычный поиск в личке
telegram_app.add_handler(
    MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        search_flower
    ),
    group=2
)


# =========================================================
# RENDER HOME
# =========================================================

@app.route(
    "/",
    methods=["GET"]
)
def home():
    return "🌸 СКЛАД КУСТИКОВ работает!"


# =========================================================
# TELEGRAM WEBHOOK
# =========================================================

@app.route(
    "/telegram/" + WEBHOOK_SECRET,
    methods=["POST"]
)
async def telegram_webhook():

    data = request.get_json(
        force=True
    )

    if not data:
        return "OK"

    update = Update.de_json(
        data,
        telegram_app.bot
    )

    await telegram_app.initialize()

    try:

        if update.effective_user:

            save_user(
                update.effective_user,
                update.effective_chat.id
                if update.effective_chat
                else None
            )

        await telegram_app.process_update(
            update
        )

    except Exception as e:

        print(
            "WEBHOOK ERROR:",
            repr(e)
        )

    finally:

        await telegram_app.shutdown()

    return "OK"


# =========================================================
# STARTUP
# =========================================================

print("BOT: запуск приложения...")

init_db()

import_excel()

print("BOT: приложение готово.")
